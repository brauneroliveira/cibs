import xml.etree.cElementTree as ET
import pandas
import csv
from openpyxl import load_workbook
from random import randint
from datetime import datetime
from xml.etree import ElementTree
from xml.dom import minidom

wb = load_workbook('bd.xlsx')
cnes = wb["CNES"]
tipo = wb["Tipo"]
qtdLeitos = 50
qtdArquivos = 1000

for i in range(qtdArquivos):
	leito = ET.Element("leito", id=str(randint(1,qtdLeitos)))
	header = ET.SubElement(leito, "header")

	ET.SubElement(header, "nivel").text = "2"
	ET.SubElement(header, "referencia").text = "CNES_GENERICO"
	ET.SubElement(header, "cnes").text = str(cnes.cell(row=randint(1,872), column=1).value)
	ET.SubElement(header, "date").text = str(datetime.now())

	ET.SubElement(leito, "area").text = "area/subarea/subsubarea..."
	ET.SubElement(leito, "tipo").text = str(tipo.cell(row=randint(1,7), column=1).value)
	ET.SubElement(leito, "estado").text = str(0)

	tree = ET.ElementTree(leito)
	filename = "xmlfiles/file" + str(i) + ".xml"
	tree.write(str(filename), encoding='utf-8', xml_declaration=True)